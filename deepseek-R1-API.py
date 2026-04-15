import openai
import openpyxl
import time
import logging

# 设置日志文件
logging.basicConfig(
    filename="deepseek_log.txt",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)

# 配置 API 密钥和模型
API_KEY = "your key"  # 替换为你的 DeepSeek API Key
BASE_URL = "https://api.deepseek.com"  # DeepSeek 平台 URL
MODEL = "deepseek-r1"

# 初始化 OpenAI 客户端
openai.api_key = API_KEY
openai.api_base = BASE_URL

# 配置重试机制
RETRY_LIMIT = 3
RETRY_DELAY = 5  # 秒

def fetch_answer_from_api(prompt, retry_count=0):
    """
    调用 DeepSeek API 获取答案。
    """
    try:
        response = openai.ChatCompletion.create(
            model=MODEL,
            messages=[
                {"role": "system", "content": "You are a helpful assistant."},
                {"role": "user", "content": prompt},
            ],
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        logging.error(f"API 调用失败: {e}")
        if retry_count < RETRY_LIMIT:
            time.sleep(RETRY_DELAY)
            return fetch_answer_from_api(prompt, retry_count + 1)
        else:
            logging.error(f"API 调用重试超过限制: {prompt}")
            raise

def main():
    # 打开 Excel 文件
    try:
        workbook = openpyxl.load_workbook("D:/A/2023.xlsx")
        sheet = workbook.active
    except Exception as e:
        logging.critical("无法打开 2023.xlsx 文件: %s", e)
        return

    # 读取固定提示词
    fixed_prompt = sheet["E2"].value
    if not fixed_prompt:
        logging.critical("E2 单元格为空，无法读取固定提示词")
        return

    # 遍历第一列中的问题
    for row in range(1, sheet.max_row + 1):
        question_cell = sheet.cell(row=row, column=1)
        answer_cell = sheet.cell(row=row, column=2)

        # 如果第2列中已经有答案，则跳过
        if answer_cell.value:
            logging.info(f"第 {row} 行已存在答案，跳过")
            continue

        question = question_cell.value
        if not question:
            logging.warning(f"第 {row} 行的问题为空，跳过")
            continue

        # 组合提示词
        prompt = f"{fixed_prompt} {question}"

        # 调用 API 获取答案
        try:
            answer = fetch_answer_from_api(prompt)
            answer_cell.value = answer
            workbook.save("D:/A/2023.xlsx")  # 实时保存
            logging.info(f"第 {row} 行: 问题—{question}；答案—{answer}")
            print(f"第 {row} 行: 已完成 (问题—{question}；答案—{answer})")
        except Exception as e:
            logging.error(f"第 {row} 行处理失败: {e}")
            print(f"第 {row} 行处理失败: {e}")

    # 结束处理
    print("所有问题处理完成。")
    logging.info("所有问题处理完成。")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        logging.critical("程序运行失败: %s", e)
