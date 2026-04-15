import pandas as pd
import requests
import time
from openpyxl import load_workbook

# 定义API的相关信息
API_URL = "https://api.siliconflow.cn/v1/chat/completions"  # 替换为实际的API地址
API_HEADERS = {
    "Authorization": "Bearer YOUR_API_KEY",  # 替换为实际的API密钥
    "Content-Type": "application/json"
}

# 输入文件路径
INPUT_FILE = 'D:/A/2023.xls'

def fetch_answer(prompt):
    """
    调用 DEEPSEEK-R1 API 获取答案
    :param prompt: str, 发送给 API 的完整提示词
    :return: str, 返回的答案
    """
    payload = {
        "model": "deepseek-ai/DeepSeek-R1",
        "stream": False,
        "max_tokens": 512,
        "temperature": 0.7,
        "top_p": 0.7,
        "top_k": 50,
        "frequency_penalty": 0.5,
        "n": 1,
        "stop": [],
        "messages": [{"role": "user", "content": prompt}]
    }

    try:
        response = requests.post(API_URL, json=payload, headers=API_HEADERS)
        response.raise_for_status()  # 检查HTTP状态码
        data = response.json()
        return data.get('choices', [{}])[0].get('message', {}).get('content', "API Error")
    except requests.exceptions.RequestException as e:
        print(f"Error during API call: {e}")
        return "Error"

def main():
    # Step 1: 读取 Excel 文件
    print("Reading input file...")
    workbook = load_workbook(INPUT_FILE)
    sheet = workbook.active

    # 获取固定提示词（E2 单元格内容）
    fixed_prompt = sheet['E2'].value
    if not fixed_prompt:
        print("Error: Fixed prompt (E2) is empty.")
        return

    # 获取问题列表（第 1 列内容）
    questions = [sheet.cell(row=row, column=1).value for row in range(1, 21)]
    if not any(questions):
        print("Error: No questions found in the first column.")
        return

    # Step 2: 遍历问题，逐步调用 API 并写入答案
    print("Processing questions and writing answers in real-time...")
    for i, question in enumerate(questions, start=1):
        if not question:
            print(f"Skipping empty question at row {i}.")
            continue

        # 构造完整提示词
        prompt = f"{fixed_prompt} {question}"

        # 调用 API 获取答案
        answer = fetch_answer(prompt)

        # 将答案写入文件的第 2 列
        sheet.cell(row=i, column=2, value=answer)
        workbook.save(INPUT_FILE)  # 实时保存文件

        # 打印进度
        print(f"Processed question {i}: {answer[:50]}...")  # 显示答案的前50个字符

        # 延迟以避免触发速率限制
        time.sleep(1.5)  # 根据 API 的速率限制调整延迟时间

    print("All answers have been successfully saved to the file.")

if __name__ == "__main__":
    main()
